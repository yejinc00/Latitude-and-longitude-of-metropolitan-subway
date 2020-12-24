from pprint import pprint
from openpyxl import load_workbook
from openpyxl import Workbook
from geopy.geocoders import Nominatim

# get name from excel file
load_wb = load_workbook("./station_name.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

# to search for a staion name
s_name=[]
for i in range(3, 283):
    s_name.append(load_ws.cell(i, 2).value)
    i+=1
    #print(s_name[0])

# get location
app = Nominatim(user_agent='tutorial')
location=[]
for k in range(3, 283):
    location.append(app.geocode(s_name[k-3]))
    print(k-2, location[k-3][1])
    k+=1

# get latitude and Longitude
detail=[]
for l in range(3, 283): #3, 80
    detail.append(location[l-3][1])
    l+=1

# put latitude and Longitude to excel file
write_wb=Workbook()
write_ws=write_wb.active
for p in range(0, 280): #0, 77
    write_ws.cell(p+1,2, detail[p][0])
    write_ws.cell(p+1,3, detail[p][1])
    p+=1

write_wb.save('./Lat_long_v2_2.xlsx')
