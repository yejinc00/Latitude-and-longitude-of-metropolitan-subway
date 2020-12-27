from pprint import pprint
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
import folium

map_osm=folium.Map(location=[37.5541369, 126.97088667728], zoom_start=17)

# get name from excel file
load_wb = load_workbook("./Lat_long_v2_2.xlsx", data_only=True)
load_ws = load_wb['Sheet']

# to search for a staion name
lat_value=[]
long_value=[]
for i in range(1, 280):
    lat_value.append(load_ws.cell(i, 2).value)
    long_value.append(load_ws.cell(i, 3).value)
    i+=1

# pointing
for p in range(0, 279):
    folium.Marker([lat_value[p], long_value[p]], popup=load_ws.cell(p+1, 1).value).add_to(map_osm)
    p+=1
map_osm.save('./map_test_v2.html')

